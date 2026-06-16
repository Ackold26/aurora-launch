"""Sprint B4 deliverable #3 — the analyst XLSX workbook (spec §2.3).

Build the 8-sheet Launch Forecast workbook from the real forecast fixture and
assert structure, real data in the populated sheets, on-brand header styling, and
that data-gated sheets are present-but-flagged (never silently dropped).
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from aurora_launch.reporting.context import build_report_context
from aurora_launch.reporting.render_xlsx import build_launch_forecast_xlsx
from aurora_launch.sample_bundles.report_fixture import build_sample_forecast_fixture


@pytest.fixture(scope="module")
def workbook(tmp_path_factory) -> dict:
    ctx = build_report_context(build_sample_forecast_fixture())
    out = tmp_path_factory.mktemp("xlsx") / "launch_forecast.xlsx"
    manifest = build_launch_forecast_xlsx(ctx, str(out))
    return {"path": str(out), "manifest": manifest, "wb": load_workbook(str(out))}


class TestWorkbookStructure:
    def test_eight_sheets_in_spec_order(self, workbook: dict) -> None:
        assert workbook["wb"].sheetnames == [
            "Summary", "12w forecast", "26w forecast", "52w forecast",
            "Channel decomposition", "Anchors", "Proxy summary", "Diagnostics",
        ]

    def test_summary_has_real_metrics(self, workbook: dict) -> None:
        ws = workbook["wb"]["Summary"]
        # header + 3 horizons.
        assert ws.max_row == 4
        # 12w total is a real forecast number (≈166M ₽), not a placeholder.
        assert ws["B2"].value > 1_000_000
        assert ws["B2"].number_format.endswith('"₽"')

    def test_weekly_sheets_full_horizon(self, workbook: dict) -> None:
        # 12 / 26 / 52 weekly rows + header each.
        for title, weeks in (("12w forecast", 12), ("26w forecast", 26), ("52w forecast", 52)):
            assert workbook["wb"][title].max_row == weeks + 1

    def test_proxy_sheet_includes_six_dimensions(self, workbook: dict) -> None:
        ws = workbook["wb"]["Proxy summary"]
        # 5 base params + 6 similarity dimensions + header = 12 rows.
        assert ws.max_row == 12


class TestBrandStyling:
    def test_header_uses_aurora_deep_fill(self, workbook: dict) -> None:
        ws = workbook["wb"]["Summary"]
        # Aurora Deep #0A1628 → ARGB FF0A1628 (Core add_table_sheet themed header).
        assert ws["A1"].fill.fgColor.rgb == "FF0A1628"
        assert ws.freeze_panes == "A2"


class TestConditionalFormatting:
    """spec §2.3 — conditional formatting via the Core xlsx primitives (a capability
    the earlier raw-openpyxl version lacked; gained by consuming aurora_xlsx.sheets)."""

    def test_weekly_ci_width_color_scale(self, workbook: dict) -> None:
        ws = workbook["wb"]["12w forecast"]
        # 5th column is the CI-width, carrying a 3-colour scale rule.
        assert ws.max_column == 5
        assert len(ws.conditional_formatting._cf_rules) >= 1

    def test_summary_confidence_tier_fill(self, workbook: dict) -> None:
        ws = workbook["wb"]["Summary"]
        # Confidence cells take the tier badge colour (silver → deep-40 #99ADC2),
        # not the default no-fill.
        assert ws["D2"].fill.fgColor.rgb == "FF99ADC2"


class TestChannelDecomposition:
    def test_channel_sheet_populated_from_per_channel_path(self, workbook: dict) -> None:
        ws = workbook["wb"]["Channel decomposition"]
        # header + 12 weekly periods (per-channel path now feeds it).
        assert ws.max_row == 13
        # Baseline column carries a real ruble number, not a note string.
        assert isinstance(ws["B2"].value, (int, float)) and ws["B2"].value > 0


class TestAnchorsSheet:
    def test_anchors_sheet_populated_from_context(self, workbook: dict) -> None:
        ws = workbook["wb"]["Anchors"]
        # 7 anchor params + header (recipient launch assumptions, context-enrichment).
        assert ws.max_row == 8
        keys = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert any("Размер рынка" in str(k) for k in keys)
        # First value carries the real market-size figure, not the follow-up note.
        assert "₽" in str(ws["B2"].value)


class TestDiagnosticsSheet:
    def test_diagnostics_sheet_populated_transfer_method(self, workbook: dict) -> None:
        ws = workbook["wb"]["Diagnostics"]
        # 6 transfer-method rows + header.
        assert ws.max_row == 7
        keys = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert any("прокси-постериора" in str(k) for k in keys)
        # INV-50: convergence is explicitly not-applicable, never faked as r_hat=1.0.
        vals = [str(ws.cell(row=r, column=2).value) for r in range(2, ws.max_row + 1)]
        assert any("неприменимо" in v for v in vals)


class TestPendingSheetsFlagged:
    def test_all_sheets_now_populated(self, workbook: dict) -> None:
        # Every sheet lands real data: channel-decomp (per-channel path), Anchors
        # (context-enrichment), Diagnostics (transfer-method model card).
        assert set(workbook["manifest"]["pending"]) == set()
