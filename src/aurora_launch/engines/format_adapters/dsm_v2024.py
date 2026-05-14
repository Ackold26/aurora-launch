"""DSM Group 2024 format adapter (B0.5 §4.1).

DSM Group provides syndicated retail audit data: weekly sales by SKU/brand
across pharma + FMCG channels. V2024 format introduced ISO 8601 datetime
+ semicolon CSV separator (vs V2023 comma + DD.MM.YYYY).

Per spec §4.1 — full implementation as canonical pattern.
V2023/V2025 stub adapters can subclass and override format-specific bits.
"""

from __future__ import annotations

from pathlib import Path

from aurora_launch.schemas.synthetic_corpus import FormatAdapterContract


class DsmAdapterV2024:
    """DSM Group 2024 format adapter."""

    def __init__(self) -> None:
        self._metadata = FormatAdapterContract(
            adapter_id="dsm_v2024",
            adapter_version="0.1.0",
            schema_version="2024",
            sample_files_glob=["*.dsm.xlsx", "*_dsm_2024_*.csv"],
            canonical_record_mapping={
                "Бренд": "brand_name",
                "Производитель": "manufacturer_name",
                "Дата": "period_date",
                "Продажи_упаковки": "sales_volume_packs",
                "Продажи_рубли": "sales_value_rub",
                "Доля_рынка": "market_share_pct",
                "АТХ_код": "atc_code",
            },
            detected_signatures=[
                "DSM Group 2024",
                "iso8601_dates",
                "csv_separator=;",
            ],
        )

    def detect(self, file_path: str) -> bool:
        """Detect if file matches DSM V2024 format.

        Heuristic check:
        - File extension matches glob
        - First few bytes contain semicolon separator + ISO date pattern
          (full-content sniffing minimal here, real impl scans header row)

        Audit (post-1D extended) — version-collision fix: explicitly reject
        files whose name carries a different DSM year marker (2023, 2025).
        Previously V2024 detect would fire on `data.dsm.2023.xlsx` because
        registry runs V2024 first (registered last) и it matched on the
        substring `.dsm` + `.xlsx` без ever inspecting the year.
        """
        path = Path(file_path)
        name_lower = path.name.lower()

        # Reject files explicitly tagged as another DSM year — V2023 / V2025
        # adapters own those (registered earlier). Без этого V2024 wins ties.
        for foreign_year_marker in ("2023", "2025"):
            if foreign_year_marker in name_lower and "dsm" in name_lower:
                return False

        if path.suffix.lower() in (".xlsx", ".xls") and ".dsm" in name_lower:
            return True
        if path.suffix.lower() == ".csv" and "dsm_2024" in name_lower:
            return True

        # Header sniff (only if file exists and is small enough)
        if path.exists() and path.suffix.lower() == ".csv":
            try:
                with path.open("r", encoding="utf-8-sig") as f:
                    header = f.readline()
                # V2024 signature: semicolon separator + Russian column names.
                # V2023 also has Russian names but uses comma separator —
                # require ; explicitly to avoid V2023 collision.
                has_semicolon = ";" in header
                has_v2024_field = "Бренд" in header or "Дата" in header
                # V2023's column is "Дата_продажи" — distinct token. Avoid
                # matching it inside V2024's Дата check.
                has_v2023_marker = "Дата_продажи" in header
                if has_semicolon and has_v2024_field and not has_v2023_marker:
                    return True
            except (OSError, UnicodeDecodeError):
                pass

        return False

    def parse(self, file_path: str) -> list[dict]:
        """Parse DSM V2024 file into list of canonical records.

        Returns list of dicts с canonical field names (from
        canonical_record_mapping). Strings only — caller must coerce types
        via downstream Pydantic schema.

        Note: skeleton implementation. Full DSM XLSX parsing requires
        openpyxl + multi-sheet handling (production would expand).

        Audit (post-1D): refuses files >256 MB (`MAX_INPUT_FILE_BYTES`) to
        prevent memory exhaustion от oversized or malicious input.
        """
        from aurora_launch.engines.format_adapters import assert_file_size_ok

        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"DSM file not found: {file_path}")

        assert_file_size_ok(path)

        records: list[dict] = []

        if path.suffix.lower() == ".csv":
            records = self._parse_csv(path)
        elif path.suffix.lower() in (".xlsx", ".xls"):
            records = self._parse_xlsx(path)
        else:
            raise ValueError(f"Unsupported DSM file format: {path.suffix}")

        return records

    def _parse_csv(self, path: Path) -> list[dict]:
        """Parse DSM V2024 CSV (semicolon separator)."""
        records: list[dict] = []
        with path.open("r", encoding="utf-8-sig") as f:
            header_line = f.readline().strip()
            headers = [h.strip() for h in header_line.split(";")]

            # Map source headers → canonical field names
            mapping = self._metadata.canonical_record_mapping
            canonical_headers = [mapping.get(h, h) for h in headers]

            for line in f:
                values = [v.strip() for v in line.strip().split(";")]
                if len(values) != len(headers):
                    continue
                record = dict(zip(canonical_headers, values, strict=False))
                records.append(record)

        return records

    def _parse_xlsx(self, path: Path) -> list[dict]:
        """Parse DSM V2024 XLSX. Requires openpyxl (lazy import)."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl required for DSM XLSX parsing. "
                "Install via: uv add openpyxl"
            ) from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        mapping = self._metadata.canonical_record_mapping
        canonical_headers = [mapping.get(h, h) for h in headers]

        records: list[dict] = []
        for row in rows[1:]:
            record = dict(zip(canonical_headers, [str(v) if v is not None else "" for v in row], strict=False))
            records.append(record)

        return records

    def get_metadata(self) -> FormatAdapterContract:
        return self._metadata
