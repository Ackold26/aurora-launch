"""Launch Forecast → XLSX renderer (Sprint B4 deliverable #3, data drill-down).

The analyst-facing workbook (spec §2.3): an 8-sheet workbook for custom modeling
and due diligence. Composed from the neutral report context with openpyxl (a direct
Launch dependency); the Core `aurora_xlsx` today exposes only `build_xlsx_minimal`
(the richer multi-sheet helpers are a Core backlog item), so Launch composes the
sheets itself and styles them with the Core `AURORA_HYBRID` palette — keeping the
deliverable on-brand without waiting on a Core helper.

Three sheets are data-gated like the PPTX deck: channel decomposition (§5.3) and
diagnostics need a per-channel forecast path / posterior diagnostics the orchestrator
does not emit yet, and recipient anchors are not carried on the context contract.
Those sheets are emitted with headers + a clear pending note and reported in the
manifest — never silently omitted.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aurora_reporting.primitives import AURORA_HYBRID

_THEME = AURORA_HYBRID


def _argb(hex_color: str) -> str:
    """openpyxl wants 8-hex ARGB (FF + RRGGBB)."""
    return "FF" + hex_color.lstrip("#").upper()


_HEADER_FILL = PatternFill("solid", fgColor=_argb(_THEME.deep[0]))
_HEADER_FONT = Font(name=_THEME.body, bold=True, color=_argb(_THEME.bg_white), size=11)
_ZEBRA_FILL = PatternFill("solid", fgColor=_argb(_THEME.bg_cream))
_FOCUS_FILL = PatternFill("solid", fgColor=_argb(_THEME.gold))
_BODY_FONT = Font(name=_THEME.body, size=10, color=_argb(_THEME.text_primary))
_NOTE_FONT = Font(name=_THEME.body, size=10, italic=True, color=_argb(_THEME.text_secondary))
_RUB_FMT = '# ##0 "₽"'
_PCT_FMT = "0.0"


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for ci, htext in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=htext)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
    ws.freeze_panes = "A2"


def _write_rows(
    ws: Worksheet,
    rows: list[list[Any]],
    *,
    number_cols: dict[int, str] | None = None,
    zebra: bool = True,
    focus_row: int | None = None,
) -> None:
    number_cols = number_cols or {}
    for ri, rowdata in enumerate(rows):
        excel_row = ri + 2
        for ci, value in enumerate(rowdata, start=1):
            cell = ws.cell(row=excel_row, column=ci, value=value)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "right")
            if ci in number_cols:
                cell.number_format = number_cols[ci]
            if focus_row is not None and ri == focus_row:
                cell.fill = _FOCUS_FILL
                cell.font = Font(name=_THEME.body, size=10, bold=True, color=_argb(_THEME.text_primary))
            elif zebra and ri % 2 == 1:
                cell.fill = _ZEBRA_FILL


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _pending_sheet(ws: Worksheet, headers: list[str], note: str) -> None:
    """A structurally-present sheet whose data is gated (engine/contract follow-up).

    Honest: the columns are defined so an analyst sees the intended shape, with a
    clear note instead of fabricated numbers.
    """
    _write_header(ws, headers)
    cell = ws.cell(row=2, column=1, value=note)
    cell.font = _NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    _autosize(ws, [max(len(note) // max(len(headers), 1), 18)] + [18] * (len(headers) - 1))


# ── Sheet builders ───────────────────────────────────────────────────────────


def _sheet_summary(wb: Workbook, ctx: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, ["Горизонт", "Прогноз продаж, ₽", "95% ДИ, %", "Уверенность"])
    km = ctx["executive_summary"]["key_metrics"]
    rows = [[f"{r['period_weeks']} нед.", r["total_rub"], r["ci_pct"], r["tier_label"]] for r in km]
    _write_rows(ws, rows, number_cols={2: _RUB_FMT, 3: _PCT_FMT}, focus_row=0)
    _autosize(ws, [14, 22, 12, 22])


def _sheet_weekly(wb: Workbook, ctx: dict[str, Any], section_key: str, title: str) -> bool:
    section = ctx.get(section_key)
    ws = wb.create_sheet(title)
    if section is None:
        _pending_sheet(ws, ["Неделя", "Среднее, ₽", "Нижняя 95%, ₽", "Верхняя 95%, ₽"],
                       "Горизонт не рассчитан для этого прогноза.")
        return False
    _write_header(ws, ["Неделя", "Среднее, ₽", "Нижняя 95%, ₽", "Верхняя 95%, ₽"])
    rows = [
        [r["week"], r["mean"], r["ci_lower"], r["ci_upper"]]
        for r in section["weekly_breakdown"]
    ]
    _write_rows(ws, rows, number_cols={2: _RUB_FMT, 3: _RUB_FMT, 4: _RUB_FMT})
    _autosize(ws, [10, 18, 18, 18])
    return True


def _sheet_channel(wb: Workbook, ctx: dict[str, Any]) -> bool:
    ws = wb.create_sheet("Channel decomposition")
    # §5.3 — per-channel contribution + baseline per period (engine surfaces it now).
    section = ctx.get("forecast_12w") or ctx.get("forecast_26w") or ctx.get("forecast_52w")
    cd = (section or {}).get("channel_decomposition")
    if not cd:
        _pending_sheet(ws, ["Период", "Канал", "Вклад, ₽", "Доля, %"],
                       "Горизонт не рассчитан.")
        return False
    channel_ids = list(cd["channels"])
    headers = ["Период", "Baseline, ₽", *[f"{c}, ₽" for c in channel_ids], "Итого, ₽"]
    rows = []
    for i, period in enumerate(cd["periods"]):
        base = cd["baseline"][i]
        chan_vals = [cd["channels"][c][i] for c in channel_ids]
        rows.append([period, base, *chan_vals, base + sum(chan_vals)])
    num_cols = {c: _RUB_FMT for c in range(1, len(headers))}
    _write_header(ws, headers)
    _write_rows(ws, rows, number_cols=num_cols)
    _autosize(ws, [10] + [16] * (len(headers) - 1))
    return True


def _sheet_proxy(wb: Workbook, ctx: dict[str, Any]) -> None:
    ws = wb.create_sheet("Proxy summary")
    pq = ctx["proxy_quality"]
    _write_header(ws, ["Параметр", "Значение"])
    rows = [
        ["Прокси-бренд", pq["proxy_brand"]],
        ["Категория", pq.get("proxy_category") or "—"],
        ["Период данных", pq.get("proxy_data_period") or "—"],
        ["Итоговая близость (S)", pq["radar"]["aggregate"]],
        ["Вердикт", pq["radar"]["verdict"]],
    ]
    # Per-dimension similarity scores.
    for dim, score in pq["radar"]["dimensions"].items():
        rows.append([f"  — {dim}", score])
    _write_rows(ws, rows, zebra=True)
    _autosize(ws, [28, 40])


def _sheet_anchors(wb: Workbook, ctx: dict[str, Any]) -> bool:
    ws = wb.create_sheet("Anchors")
    # Recipient anchors are not carried on the report-context contract today.
    _pending_sheet(ws, ["Поле anchor", "Значение"],
                   "Recipient anchors не входят в контракт контекста — context-enrichment follow-up.")
    return False


def _sheet_diagnostics(wb: Workbook, ctx: dict[str, Any]) -> bool:
    ws = wb.create_sheet("Diagnostics")
    diag = ctx["methodology"].get("diagnostics")
    if diag is None:
        _pending_sheet(ws, ["Метрика", "Значение"],
                       "Posterior-диагностика (Gelman-Rubin/ESS/R²/MAPE) появится из реального постериора.")
        return False
    _write_header(ws, ["Метрика", "Значение"])
    _write_rows(ws, [[k, v] for k, v in diag.items()])
    _autosize(ws, [28, 20])
    return True


def build_launch_forecast_xlsx(context: dict[str, Any], output_path: str) -> dict[str, Any]:
    """Render the 8-sheet Launch Forecast workbook to ``output_path``.

    Returns a manifest: sheet names + which sheets are pending (data-gated).
    """
    wb = Workbook()

    pending: list[str] = []
    _sheet_summary(wb, context)
    for key, title in (("forecast_12w", "12w forecast"), ("forecast_26w", "26w forecast"),
                       ("forecast_52w", "52w forecast")):
        if not _sheet_weekly(wb, context, key, title):
            pending.append(title)
    if not _sheet_channel(wb, context):
        pending.append("Channel decomposition")
    if not _sheet_anchors(wb, context):
        pending.append("Anchors")
    _sheet_proxy(wb, context)
    if not _sheet_diagnostics(wb, context):
        pending.append("Diagnostics")

    wb.save(output_path)
    return {
        "output_path": output_path,
        "sheets": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "pending": pending,
    }


def _main() -> None:
    import json

    from aurora_launch.reporting.context import build_report_context
    from aurora_launch.sample_bundles.report_fixture import build_sample_forecast_fixture

    ctx = build_report_context(build_sample_forecast_fixture())
    manifest = build_launch_forecast_xlsx(ctx, "launch_forecast_sample.xlsx")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    _main()
